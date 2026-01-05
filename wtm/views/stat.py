import openpyxl
from openpyxl.styles import Alignment,  Font
from urllib.parse import quote

from datetime import date
from calendar import monthrange

from django.contrib.auth.decorators import login_required

from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from common import context_processors
from wtm.services.attendance import build_monthly_attendance_summary_for_users, build_monthly_metric_details_for_users
from .helpers import sec_to_hhmmss, fetch_base_users_for_month, get_non_business_days
from .helpers_excel import header_fill, header_font, header_align, set_table_border, metric_excel_data, write_metric_sheet, schedule_excel_data, write_schedule_sheet


def build_work_status_rows(stand_ym: str | None):
    """
    근태기록-월간집계(전체)에서 사용하는 rows 공통 빌더.
    """
    stand_ym = stand_ym or timezone.now().strftime("%Y%m")
    year, month = int(stand_ym[:4]), int(stand_ym[4:6])

    base_users = fetch_base_users_for_month(stand_ym)
    if not base_users:
        return stand_ym, []

    uid_list = [u["user_id"] for u in base_users]
    # 퇴사일자 이후의 근무표는 무시하기 위해 해당 정보를 전달
    out_ymd_map = {u["user_id"]: u.get("out_ymd") for u in base_users}

    summary_map = build_monthly_attendance_summary_for_users(
        users=uid_list,
        year=year,
        month=month,
        out_ymd_map=out_ymd_map,
    )

    rows = []
    for u in base_users:
        s = summary_map.get(u["user_id"], {})
        rows.append({
            "dept": u["dept"],
            "position": u.get("position"),
            "emp_name": u["emp_name"],

            "error_cnt": s.get("error_count", 0),

            "reg_late_cnt": s.get("reg_late_count", 0),
            "reg_late_sec": s.get("reg_late_seconds", 0),
            "reg_early_cnt": s.get("reg_early_count", 0),
            "reg_early_sec": s.get("reg_early_seconds", 0),
            "reg_overtime_cnt": s.get("reg_overtime_count", 0),
            "reg_overtime_sec": s.get("reg_overtime_seconds", 0),

            "total_sec": s.get("hol_total_seconds", 0),

            "hol_work_cnt": s.get("hol_work_count", 0),
            "hol_work_sec": s.get("hol_work_seconds", 0),
            "hol_late_cnt": s.get("hol_late_count", 0),
            "hol_late_sec": s.get("hol_late_seconds", 0),
            "hol_early_cnt": s.get("hol_early_count", 0),
            "hol_early_sec": s.get("hol_early_seconds", 0),
            "hol_overtime_cnt": s.get("hol_overtime_count", 0),
            "hol_overtime_sec": s.get("hol_overtime_seconds", 0),

            "nopay_cnt": s.get("nopay_count", 0),
        })

    return stand_ym, rows


# 근태기록-월간집계(전체)
@login_required(login_url="common:login")
def work_status(request, stand_ym: str | None = None):
    stand_ym, rows = build_work_status_rows(stand_ym)

    return render(request, "wtm/work_status.html", {
        "stand_ym": stand_ym,
        "rows": rows,
        "active_metric": "all",
    })


@login_required(login_url="common:login")
def work_status_excel(request, stand_ym: str | None = None):
    """
    근태기록-월간집계 엑셀 다운로드 (탭 통합)
    - 어떤 탭에서 내려받아도 동일한 파일을 생성한다.
    - 시트 구성:
        1) 근태기록 : 기존 월간집계(전체)
        2) 근무표   : schedule.py / work_schedule.html 표
        3) 지각
        4) 조퇴
        5) 연장근로
        6) 휴일근로
    """
    stand_ym, rows = build_work_status_rows(stand_ym)
    year, month = int(stand_ym[:4]), int(stand_ym[4:6])

    wb = openpyxl.Workbook()

    # Sheet 1: 근태기록(기존 포맷)
    ws = wb.active
    ws.title = "근태기록"

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    ws["A1"] = f"{year}년 {month}월 근태기록 월간집계(전체)"
    ws["A1"].font = Font(size=10, bold=True)
    ws.merge_cells("A1:S1")  # 19열(A~S)

    ws["A2"] = "부서"
    ws["B2"] = "성명"
    ws["C2"] = "오류\n(일수)"
    ws["D2"] = "소정근로"
    ws["J2"] = "휴일근로"
    ws["S2"] = "무급휴무"

    ws.merge_cells("A2:A4")
    ws.merge_cells("B2:B4")
    ws.merge_cells("C2:C4")
    ws.merge_cells("D2:I2")
    ws.merge_cells("J2:R2")
    ws.merge_cells("S2:S4")

    ws["D3"] = "지각";     ws.merge_cells("D3:E3")
    ws["F3"] = "조퇴";     ws.merge_cells("F3:G3")
    ws["H3"] = "연장근로"; ws.merge_cells("H3:I3")

    ws["J3"] = "TOTAL\n(시간)"
    ws.merge_cells("J3:J4")

    ws["K3"] = "근로";     ws.merge_cells("K3:L3")
    ws["M3"] = "지각";     ws.merge_cells("M3:N3")
    ws["O3"] = "조퇴";     ws.merge_cells("O3:P3")
    ws["Q3"] = "연장근로"; ws.merge_cells("Q3:R3")

    ws["D4"] = "횟수"; ws["E4"] = "시간"
    ws["F4"] = "횟수"; ws["G4"] = "시간"
    ws["H4"] = "횟수"; ws["I4"] = "시간"

    ws["K4"] = "횟수"; ws["L4"] = "시간"
    ws["M4"] = "횟수"; ws["N4"] = "시간"
    ws["O4"] = "횟수"; ws["P4"] = "시간"
    ws["Q4"] = "횟수"; ws["R4"] = "시간"

    for rr in range(2, 5):
        ws.row_dimensions[rr].height = 24
        for cell in ws[rr]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

    def _t(sec: int) -> str:
        return "" if not sec else sec_to_hhmmss(sec)

    for r in rows:
        ws.append([
            r.get("dept") or "",
            r.get("emp_name") or "",
            r.get("error_cnt") or "",

            r.get("reg_late_cnt") or "", _t(r.get("reg_late_sec") or 0),
            r.get("reg_early_cnt") or "", _t(r.get("reg_early_sec") or 0),
            r.get("reg_overtime_cnt") or "", _t(r.get("reg_overtime_sec") or 0),

            _t(r.get("total_sec") or 0),

            r.get("hol_work_cnt") or "", _t(r.get("hol_work_sec") or 0),
            r.get("hol_late_cnt") or "", _t(r.get("hol_late_sec") or 0),
            r.get("hol_early_cnt") or "", _t(r.get("hol_early_sec") or 0),
            r.get("hol_overtime_cnt") or "", _t(r.get("hol_overtime_sec") or 0),

            r.get("nopay_cnt") or "",
        ])

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=19):
        for cell in row:
            cell.font = Font(size=10)
            if cell.column in (1, 2):
                cell.alignment = align_center
            else:
                cell.alignment = align_right

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10

    ws.column_dimensions["D"].width = 8;  ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8;  ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 8;  ws.column_dimensions["I"].width = 12

    ws.column_dimensions["J"].width = 14

    ws.column_dimensions["K"].width = 8;  ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 8;  ws.column_dimensions["N"].width = 12
    ws.column_dimensions["O"].width = 8;  ws.column_dimensions["P"].width = 12
    ws.column_dimensions["Q"].width = 8;  ws.column_dimensions["R"].width = 12

    ws.column_dimensions["S"].width = 10

    set_table_border(ws, 2, ws.max_row, 1, 19)

    # Sheet 2: 근무표
    ws_schedule = wb.create_sheet(title="근무표")
    schedule_data = schedule_excel_data(stand_ym)
    write_schedule_sheet(ws_schedule, schedule_data)

    # Sheets 3~6: 지각/조퇴/연장근로/휴일근로
    metric_specs = [
        ("late", "지각"),
        ("early", "조퇴"),
        ("overtime", "연장근로"),
        ("holiday", "휴일근로"),
    ]
    for metric, title in metric_specs:
        ws_metric = wb.create_sheet(title=title)
        metric_data = metric_excel_data(stand_ym, metric)
        write_metric_sheet(ws_metric, metric_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = quote(f"근태기록-월간집계_{stand_ym}.xlsx")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# 근태기록-월간집계(지각/조퇴/연장/휴일근로)
@login_required(login_url="common:login")
def work_metric(request, metric: str, stand_ym: str | None = None):
    """
    지각/조퇴/연장/휴근 탭 화면 (월간 일자 그리드).
    - unit: month
    - calc: seconds (템플릿에서 HH:MM:SS 포맷)
    """
    _ALLOWED_METRICS = {"late": "지각", "early": "조퇴", "overtime": "연장근로", "holiday": "휴일근로"}
    metric = metric.lower()
    if metric not in _ALLOWED_METRICS:
        return render(request, "wtm/work_metric.html", {
            "error": "잘못된 지표 요청입니다.",
        })

    stand_ym = stand_ym or timezone.now().strftime("%Y%m")
    year, month = int(stand_ym[:4]), int(stand_ym[4:6])

    # 1) 대상자 공통 헬퍼
    base_users = fetch_base_users_for_month(stand_ym)

    if not base_users:
        return render(request, "wtm/work_metric.html", {
            "metric": metric,
            "metric_label": _ALLOWED_METRICS[metric],
            "stand_ym": stand_ym,
            "rows": [],
            "day_list": {}, "holiday_list": [],
            "active_metric": "all",
        })

    # 2) 일자 리스트 (헤더용)
    day_list = context_processors.get_day_list(stand_ym)
    holiday_list = sorted(get_non_business_days(int(stand_ym[0:4]), int(stand_ym[4:6])))

    # 3) 메트릭 디테일 계산
    uid_list = [u["user_id"] for u in base_users]
    detail_map = build_monthly_metric_details_for_users(users=uid_list, year=year, month=month, metric=metric)

    # 4) 템플릿 rows 구성 (좌측: 부서/직급/성명/횟수/합계, 오른쪽: 일자별 초)
    days_order = list(range(1, monthrange(year, month)[1] + 1))
    rows = []
    for u in base_users:
        d = detail_map.get(u["user_id"], {"days": {}, "count": 0, "total_seconds": 0})
        cell_seconds = [d["days"].get(date(year, month, dd), 0) for dd in days_order]

        day_secs = []
        for dd, sec in zip(days_order, cell_seconds):
            if sec and sec > 0:
                dow = day_list.get(dd) or day_list.get(str(dd)) or ""
                ymd = f"{year:04d}{month:02d}{dd:02d}"
                day_secs.append({
                    "day": dd,
                    "dow": dow,
                    "sec": sec,
                    "ymd": ymd,
                })

        rows.append({
            "user_id": u["user_id"],
            "dept": u["dept"],
            "position": u.get("position"),
            "emp_name": u["emp_name"],
            "count": d["count"],
            "total_seconds": d["total_seconds"],
            "cells": cell_seconds,
            "day_secs": day_secs,
        })

    return render(request, "wtm/work_metric.html", {
        "metric": metric,
        "metric_label": _ALLOWED_METRICS[metric],
        "stand_ym": stand_ym,
        "day_list": day_list,
        "holiday_list": holiday_list,
        "rows": rows,
        "active_metric": metric,
    })
