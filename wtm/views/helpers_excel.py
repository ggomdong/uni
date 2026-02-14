from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from datetime import date
from calendar import monthrange

from django.db.models.functions import ExtractDay
from django.db import connection

from common import context_processors
from common.models import Holiday
from wtm.services.base_users import fetch_base_users_for_month
from ..services.attendance import build_monthly_metric_details_for_users
from ..models import Module, Schedule
from .helpers import sec_to_hhmmss, get_non_business_days


EXCEL_HEADER_BG = "FFA8E5E8"     # #A8E5E8
EXCEL_HEADER_FONT = "FF000000"  # 검정
EXCEL_HEADER_RED = "FFFF0000"   # 빨강

header_fill = PatternFill(fill_type="solid", start_color=EXCEL_HEADER_BG, end_color=EXCEL_HEADER_BG)
header_font = Font(size=10, bold=True, color=EXCEL_HEADER_FONT)
header_font_red = Font(size=10, bold=True, color=EXCEL_HEADER_RED)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

body_font = Font(size=10)
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")


def set_table_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border


def metric_excel_data(stand_ym: str, metric: str, *, branch) -> dict:
    """
    work_metric 화면과 동일한 데이터(월 그리드)를 엑셀용으로 구성.
    반환:
        {
            "metric": "late|early|overtime|holiday",
            "metric_label": "지각|조퇴|연장근로|휴일근로",
            "stand_ym": "YYYYMM",
            "year": int,
            "month": int,
            "day_list": dict,
            "rows": [
                {"dept","emp_name","count","total_seconds","cells":[sec,...]},
            ]
        }
    """
    _ALLOWED_METRICS = {"late": "지각", "early": "조퇴", "overtime": "연장근로", "holiday": "휴일근로"}
    metric = metric.lower()
    if metric not in _ALLOWED_METRICS:
        raise ValueError("invalid metric")

    year, month = int(stand_ym[:4]), int(stand_ym[4:6])
    base_users = fetch_base_users_for_month(stand_ym, branch=branch)
    day_list = context_processors.get_day_list(stand_ym)

    if not base_users:
        return {
            "metric": metric,
            "metric_label": _ALLOWED_METRICS[metric],
            "stand_ym": stand_ym,
            "year": year,
            "month": month,
            "day_list": day_list,
            "rows": [],
        }

    uid_list = [u["user_id"] for u in base_users]
    detail_map = build_monthly_metric_details_for_users(
        users=uid_list,
        year=year,
        month=month,
        metric=metric,
        branch=branch,
    )

    days_order = list(range(1, monthrange(year, month)[1] + 1))
    rows = []
    for u in base_users:
        d = detail_map.get(u["user_id"], {"days": {}, "count": 0, "total_seconds": 0})
        cell_seconds = [d["days"].get(date(year, month, dd), 0) for dd in days_order]
        rows.append({
            "dept": u["dept"],
            "emp_name": u["emp_name"],
            "count": d.get("count", 0),
            "total_seconds": d.get("total_seconds", 0),
            "cells": cell_seconds,
        })

    non_business_days = get_non_business_days(year, month, branch=branch)

    return {
        "metric": metric,
        "metric_label": _ALLOWED_METRICS[metric],
        "stand_ym": stand_ym,
        "year": year,
        "month": month,
        "day_list": day_list,
        "rows": rows,
        "non_business_days": non_business_days,
    }


def write_metric_sheet(ws, data: dict) -> None:
    """
    지각/조퇴/연장근로/휴일근로 시트 작성.
    - 1행: (A~D) 지표명 병합 + (E~) 일자 숫자
    - 2행: 부서/성명/횟수/합계 + 요일
    - 3행~: 데이터(일자별 HH:MM:SS)
    """
    year, month = data["year"], data["month"]
    day_list = data["day_list"]
    non_business_days = data.get("non_business_days", set())
    last_day = monthrange(year, month)[1]

    def _t(sec: int) -> str:
        return "" if not sec else sec_to_hhmmss(sec)

    # 1행: (A~D) 지표명 병합 + (E~) 일자
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=1, column=1, value=data["metric_label"]).font = header_font
    ws.cell(row=1, column=1).alignment = header_align

    for dd in range(1, last_day + 1):
        col = 4 + dd
        c = ws.cell(row=1, column=col, value=dd)
        c.font = header_font_red if (dd in non_business_days) else header_font
        c.alignment = header_align

    # 2행: 고정 헤더 + 요일
    fixed_headers = ["부서", "성명", "횟수", "합계"]
    for idx, h in enumerate(fixed_headers, start=1):
        c = ws.cell(row=2, column=idx, value=h)
        c.font = header_font
        c.alignment = header_align

    for dd in range(1, last_day + 1):
        col = 4 + dd
        dow = day_list.get(dd) or day_list.get(str(dd)) or ""
        c = ws.cell(row=2, column=col, value=dow)

        c.font = header_font_red if (dd in non_business_days) else header_font
        c.alignment = header_align

    # 헤더(1~2행) 스타일 일괄 적용: Fill
    for rr in (1, 2):
        ws.row_dimensions[rr].height = 24
        for cell in ws[rr]:
            cell.fill = header_fill

    # 3행~ 데이터
    for r in data["rows"]:
        row_idx = ws.max_row + 1

        c = ws.cell(row=row_idx, column=1, value=r.get("dept") or "")
        c.font = body_font
        c.alignment = align_center

        c = ws.cell(row=row_idx, column=2, value=r.get("emp_name") or "")
        c.font = body_font
        c.alignment = align_center

        c = ws.cell(row=row_idx, column=3, value=r.get("count") or "")
        c.font = body_font
        c.alignment = align_right

        c = ws.cell(row=row_idx, column=4, value=_t(r.get("total_seconds") or 0))
        c.font = body_font
        c.alignment = align_right

        cells = r.get("cells") or []
        for dd in range(1, last_day + 1):
            sec = cells[dd - 1] if dd - 1 < len(cells) else 0
            c = ws.cell(row=row_idx, column=4 + dd, value=_t(sec))
            c.font = body_font
            c.alignment = align_center

    # 컬럼 폭/고정
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 12
    for dd in range(1, last_day + 1):
        ws.column_dimensions[get_column_letter(4 + dd)].width = 8

    ws.freeze_panes = "E3"
    set_table_border(ws, 1, ws.max_row, 1, 4 + last_day)


def schedule_excel_data(stand_ym: str, *, branch) -> dict:
    """
    schedule.py / work_schedule.html 표를 엑셀용으로 구성
    """
    year, month = int(stand_ym[:4]), int(stand_ym[4:6])
    last_day = monthrange(year, month)[1]

    day_list = context_processors.get_day_list(stand_ym)

    schedule_count = Schedule.objects.filter(
        year=stand_ym[:4],
        month=stand_ym[4:6],
        branch=branch,
    ).count()
    if schedule_count == 0:
        return {
            "stand_ym": stand_ym,
            "year": year,
            "month": month,
            "day_list": day_list,
            "rows": [],
            "has_schedule": False,
        }

    schedule_date = f"{stand_ym}{last_day:02d}"

    query_user = '''
                SELECT u.id, u.emp_name, u.dept, u.position,
                        DATE_FORMAT(u.join_date, '%%Y%%m%%d') join_date, DATE_FORMAT(u.out_date, '%%Y%%m%%d') out_date,
                        d.`order` as do, p.`order` as po
                FROM common_user u
                    LEFT OUTER JOIN common_dept d on (u.dept = d.dept_name AND d.branch_id = u.branch_id)
                    LEFT OUTER JOIN common_position p on (u.position = p.position_name AND p.branch_id = u.branch_id)
                WHERE is_employee = TRUE
                    AND u.branch_id = %s
                    and DATE_FORMAT(u.join_date, '%%Y%%m%%d') <= %s
                    and (DATE_FORMAT(u.out_date, '%%Y%%m%%d') is null or DATE_FORMAT(u.out_date, '%%Y%%m%%d') >= %s)
                ORDER BY do, po, join_date, emp_name
                '''

    with connection.cursor() as cursor:
        cursor.execute(query_user, [branch.id, schedule_date, f"{stand_ym}01"])
        results = cursor.fetchall()
        desc = cursor.description
        base_users = []
        for r in results:
            d = {}
            for i in range(len(desc)):
                d[desc[i][0]] = r[i]
            base_users.append(d)

    if not base_users:
        return {
            "stand_ym": stand_ym,
            "year": year,
            "month": month,
            "day_list": day_list,
            "rows": [],
            "has_schedule": True,
        }

    user_ids = [u["id"] for u in base_users]

    sched_map = {
        s.user_id: s
        for s in Schedule.objects.filter(
            year=stand_ym[:4],
            month=stand_ym[4:6],
            user_id__in=user_ids,
            branch=branch,
        )
    }

    module_map = {m.id: {"cat": m.cat, "name": m.name} for m in Module.objects.filter(branch=branch)}

    rows = []
    for u in base_users:
        user_id = u["id"]
        join_date = u.get("join_date") or ""
        out_date = u.get("out_date") or ""

        sched = sched_map.get(user_id)

        day_values = []
        workday_cnt = 0

        for dd in range(1, last_day + 1):
            ymd = f"{stand_ym}{dd:02d}"

            if join_date and ymd < join_date:
                day_values.append("")
                continue
            if out_date and out_date != "" and ymd > out_date:
                day_values.append("")
                continue

            module_id = getattr(sched, f"d{dd}_id", None) if sched else None
            if not module_id:
                day_values.append("")
                continue

            m = module_map.get(module_id)
            if not m:
                day_values.append("")
                continue

            if m["cat"] == "OFF":
                day_values.append("OFF")
            else:
                day_values.append(m["name"])
                workday_cnt += 1

        rows.append({
            "dept": u["dept"],
            "emp_name": u["emp_name"],
            "count": workday_cnt,
            "total": "-",
            "days": day_values,
        })

    non_business_days = get_non_business_days(year, month, branch=branch)

    return {
        "stand_ym": stand_ym,
        "year": year,
        "month": month,
        "day_list": day_list,
        "rows": rows,
        "has_schedule": True,
        "non_business_days": non_business_days,
    }


def write_schedule_sheet(ws, data: dict) -> None:
    """
    근무표 시트 작성(첨부2 형태)
    - 1행: (A~D) '근무표' 병합 + (E~) 일자 숫자
    - 2행: 부서/성명/횟수/합계 + 요일
    """
    year, month = data["year"], data["month"]
    day_list = data["day_list"]
    non_business_days = data.get("non_business_days", set())
    last_day = monthrange(year, month)[1]

    # 1행: 근무표 + 일자
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=1, column=1, value="근무표").font = header_font
    ws.cell(row=1, column=1).alignment = header_align

    for dd in range(1, last_day + 1):
        col = 4 + dd
        c = ws.cell(row=1, column=col, value=dd)
        c.font = header_font_red if (dd in non_business_days) else header_font
        c.alignment = header_align

    # 2행: 부서/성명/횟수/합계 + 요일
    fixed_headers = ["부서", "성명", "횟수", "합계"]
    for idx, h in enumerate(fixed_headers, start=1):
        c = ws.cell(row=2, column=idx, value=h)
        c.font = header_font
        c.alignment = header_align

    for dd in range(1, last_day + 1):
        col = 4 + dd
        dow = day_list.get(dd) or day_list.get(str(dd)) or ""
        c = ws.cell(row=2, column=col, value=dow)

        c.font = header_font_red if (dd in non_business_days) else header_font
        c.alignment = header_align

    # 헤더(1~2행) Fill
    for rr in (1, 2):
        ws.row_dimensions[rr].height = 24
        for cell in ws[rr]:
            cell.fill = header_fill

    # 데이터
    if not data.get("has_schedule"):
        c = ws.cell(row=3, column=1, value="근무표가 없습니다.")
        c.font = body_font
        c.alignment = align_center
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4 + last_day)
    else:
        for r in data["rows"]:
            row_idx = ws.max_row + 1

            c = ws.cell(row=row_idx, column=1, value=r.get("dept") or "")
            c.font = body_font
            c.alignment = align_center

            c = ws.cell(row=row_idx, column=2, value=r.get("emp_name") or "")
            c.font = body_font
            c.alignment = align_center

            c = ws.cell(row=row_idx, column=3, value=r.get("count") or "")
            c.font = body_font
            c.alignment = align_right

            c = ws.cell(row=row_idx, column=4, value=r.get("total") or "-")
            c.font = body_font
            c.alignment = align_center

            days = r.get("days") or []
            for dd in range(1, last_day + 1):
                v = days[dd - 1] if dd - 1 < len(days) else ""
                c = ws.cell(row=row_idx, column=4 + dd, value=v)
                c.font = body_font
                c.alignment = align_center

    # 폭/고정/테두리
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 8
    for dd in range(1, last_day + 1):
        ws.column_dimensions[get_column_letter(4 + dd)].width = 8

    ws.freeze_panes = "E3"
    set_table_border(ws, 1, ws.max_row, 1, 4 + last_day)
